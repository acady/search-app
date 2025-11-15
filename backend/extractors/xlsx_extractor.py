from io import BytesIO
from openpyxl import load_workbook

def extract_text_from_xlsx(file_bytes: bytes) -> str:
    bio = BytesIO(file_bytes)
    wb = load_workbook(bio, data_only=True)
    texts = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    texts.append(str(cell))
    return "\n".join(texts)
